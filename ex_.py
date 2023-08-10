# import configparser

# def create_config_file():
#     config = configparser.ConfigParser()

#     config.add_section('Powersupply Test')
#     powersupply_values = [
#         "DCV b/w GND - R709",
#         "DCV b/w GND - R700",
#         "ACV b/w GND - R709",
#         "ACV b/w GND - R700",
#         "DCV b/w ISOGND - C443",
#         "DCV b/w ISOGND - C442",
#         "DCV b/w ISOGND - C441",
#         "DCV b/w ISOGND - C412",
#         "DCV 2.048V b/w ISOGND - C430",
#         "ACV b/w ISOGND - C443",
#         "ACV b/w ISOGND - C442",
#         "ACV b/w ISOGND - C441",
#         "ACV b/w ISOGND - C412"
#     ]
#     for value in powersupply_values:
#         config.set('Powersupply Test', value, '')

#     config.add_section('I2C Test')
#     i2c_values = [
#         "UID",
#         "IC704 result registers Reading"
#     ]
#     for value in i2c_values:
#         config.set('I2C Test', value, '')

#     with open('conf_ig.ini', 'w') as configfile:
#         config.write(configfile)

# def update_value_in_section(section, key, new_value):
#     config = configparser.ConfigParser()
#     config.read('conf_ig.ini')

#     if config.has_section(section) and config.has_option(section, key):
#         config.set(section, key, str(new_value))  # Convert float to string

#         with open('conf_ig.ini', 'w') as configfile:
#             config.write(configfile)

# if __name__ == "__main__":
#     create_config_file()
#     print("conf_ig.ini file created successfully.")

#     update_value_in_section('Powersupply Test', 'DCV b/w GND - R709', 53.5)
#     print("Value updated in 'DCV b/w GND - R709' successfully.")


# import configparser
# import openpyxl

# def convert_ini_to_excel(ini_file, excel_file):
#     config = configparser.ConfigParser()
#     config.read(ini_file)

#     workbook = openpyxl.Workbook()
#     for section in config.sections():
#         worksheet = workbook.create_sheet(title=section)
#         for key, value in config.items(section):
#             worksheet.append([key, value])

#     workbook.remove(workbook['Sheet'])  # Remove default sheet
#     workbook.save(excel_file)

# if __name__ == "__main__":
#     ini_file = 'conf_ig.ini'
#     excel_file = 'conf_ig.xlsx'

#     convert_ini_to_excel(ini_file, excel_file)
#     print(f"INI file '{ini_file}' converted to Excel file '{excel_file}' successfully.")





# import configparser

# def print_ini_file(file_path):
#     config = configparser.ConfigParser()
#     config.read(file_path)

#     for section in config.sections():
#         print(f"[{section}]")
#         for key, value in config.items(section):
#             print(f"{key} = {value}")
#         print()

# if __name__ == "__main__":
#     ini_file = 'conf_ig.ini'
#     print_ini_file(ini_file)


# import configparser
# import openpyxl

# def convert_ini_to_excel(ini_file, excel_file):
#     config = configparser.ConfigParser()
#     config.read(ini_file)

#     workbook = openpyxl.Workbook()

#     for section in config.sections():
#         worksheet = workbook.create_sheet(title=section)
#         worksheet.append(["Section", "Key", "Value"])  # Add headers

#         for key, value in config.items(section):
#             worksheet.append([section, key, value])

#     workbook.remove(workbook['Sheet'])  # Remove default sheet
#     workbook.save(excel_file)

# if __name__ == "__main__":
#     ini_file = 'conf_ig.ini'
#     excel_file = 'conf_ig.xlsx'

#     convert_ini_to_excel(ini_file, excel_file)
#     print(f"INI file '{ini_file}' converted to Excel file '{excel_file}' successfully.")



import configparser
import openpyxl

def convert_ini_to_excel(ini_file, excel_file):
    config = configparser.ConfigParser()
    config.read(ini_file)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'All Sections'

    worksheet.append(["Section", "Key", "Value"])  # Add headers

    for section in config.sections():
        for key, value in config.items(section):
            worksheet.append([section, key, value])

    workbook.save(excel_file)

if __name__ == "__main__":
    ini_file = 'conf_ig.ini'
    excel_file = 'conf_ig.xlsx'

    convert_ini_to_excel(ini_file, excel_file)
    print(f"INI file '{ini_file}' converted to Excel file '{excel_file}' successfully.")