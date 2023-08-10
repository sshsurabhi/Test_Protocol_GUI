# import openpyxl

# def read_excel(file_path, sheet_name):
#     try:
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook[sheet_name]
        
#         for row in sheet.iter_rows(values_only=True):
#             for cell in row:
#                 print(cell, end='\t')
#             print()
            
#     except Exception as e:
#         print(f"An error occurred: {e}")

# if __name__ == "__main__":
#     excel_file = "configs/Test_result.xlsx"
#     sheet_name = "Sheet1"  # Change this to the name of your sheet
    
#     read_excel(excel_file, sheet_name)

import openpyxl

def modify_excel(file_path, sheet_name, modifications):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        for row_index, (cell_value, new_value) in enumerate(modifications, start=2):
            sheet.cell(row=row_index, column=2, value=new_value)
        
        workbook.save(file_path)
        print("Modifications saved successfully.")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    excel_file = "configs/Test_result.xlsx"
    sheet_name = "Sheet1"  # Change this to the name of your sheet
    
    modifications = [
        ("Voltage Set", 1),
        ("Current Set", 2),
        ("Read Current before Soldering", 3),
        ("Voltage read before Jumper Close", 4),
        ("Current read after Jumper close", 5),
        ("DCV b/w GND - R709", 6),
        ("DCV b/w GND - R700", 7),
        ("ACV b/w GND - R709", 8),
        ("ACV b/w GND - R700", 9),
        ("DCV b/w ISOGND - C443", 10),
        ("DCV b/w ISOGND - C442", 11),
        ("DCV b/w ISOGND - C441", 12),
        ("DCV b/w ISOGND - C412", 13),
        ("DCV 2.048V b/w ISOGND - C430", 14),
        ("ACV b/w ISOGND - C443", 15),
        ("ACV b/w ISOGND - C442", 16),
        ("ACV b/w ISOGND - C441", 17),
        ("ACV b/w ISOGND - C412", 18),
        ("ACV 2.048 b/w ISOGND - C430", 19),
        ("Current Read after SDIO board Fix", 20),
        ("Current Read after FPGA & SD-Card Fix", 21),
        ("UID", 22),
        ("IC704 result registers Reading", 23),
    ]
    
    modify_excel(excel_file, sheet_name, modifications)
