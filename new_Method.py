import datetime
import os
import sys
import time

import openpyxl
import pandas as pd
import pyvisa as visa
import serial
from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *


class WorkerThread(QThread):
    process_completed = pyqtSignal()
    result_signal = pyqtSignal(str)
    response_signal = pyqtSignal(str)
    def __init__(self, commands, serial_port):
        super().__init__()
        self.commands = commands
        self.serial_port = serial_port
    def run(self):
        if self.serial_port.is_open:
            self.serial_port.timeout = 5
            self.next_command_index = 0
            for command in self.commands:
                if self.next_command_index < (len(self.commands)+1):
                    self.serial_port.write(command.encode())
                    # self.textBrowser.append('Processed Command: '+command + ' ')
                    response = self.serial_port.readline().decode('ascii')
                    self.result_signal.emit((f"Response: {response}"))
                    self.next_command_index += 1
                    current_date = datetime.datetime.now()
                    decimal_date = int(current_date.strftime('%Y%m%d'))
                    hex_date = hex(decimal_date)[2:].upper().zfill(8)
                    if command == self.commands[1]:
                        ading = response.split(':')[3][:-1]
                        self.commands[2] = self.commands[2]+ading+'01'+hex_date+'2A0101030000FFFF'
                        self.response_signal.emit(ading)
                    if command == self.commands[2]:
                        start_time = time.time()
                        while time.time() - start_time < 5:
                            if response:
                                break
                            response = self.serial_port.readline().decode('ascii')
                            self.result_signal.emit(response)
            self.serial_port.close()
            self.process_completed.emit()
        else:
            self.result_signal.emit('Serial Port Closed')
    def on_button_clicked(self):
        self.result_signal.emit("Button clicked")

class SerialPortThread(QThread):        #    Thread to update com ports in the network
    com_ports_available = pyqtSignal(list)
    def run(self):
        com_ports = []
        for i in range(256):
            try:
                s = serial.Serial('COM'+str(i))
                com_ports.append('COM'+str(i))
                s.close()
            except serial.SerialException:
                pass
        self.com_ports_available.emit(com_ports)

class App(QMainWindow):
    def __init__(self):
        super(App, self).__init__()
        uic.loadUi("UI/ascha.ui", self)
        self.setWindowIcon(QIcon('images_/icons/2.jpg'))
        self.setFixedSize(self.size())
        self.setStatusTip('Moewe Optik Gmbh. 2023')
        self.show()
#############################################################################################################################################################

        self.serial_port = None
        self.thread = None
        self.serial_thread = SerialPortThread()
        self.serial_thread.com_ports_available.connect(self.update_com_ports)
        self.serial_thread.start()
        self.baudrate_box.addItems(['9600','57600','115200'])
        self.baudrate_box.setCurrentText('115200')
        self.connect_button.clicked.connect(self.connect_or_disconnect_serial_port)
        self.refresh_button.clicked.connect(self.refresh_connect)
###################################################################################################
        self.start_button.clicked.connect(self.connect)
        # self.timer = QTimer(self)
        # self.timer.timeout.connect(self.update_time_label)
        # self.timer.start(1000) 
        self.rm = visa.ResourceManager()
        self.multimeter = None
        self.powersupply = None
        self.AC_DC_box.addItems(['<select>', 'DCV', 'ACV'])
        self.AC_DC_box.currentTextChanged.connect(self.handleSelectionChange)
        self.test_button.clicked.connect(self.on_cal_voltage_current)




        self.DC_values = []
        self.AC_values = []
        self.PS_channel = None
        self.max_voltage = 0
        self.max_current = 0
        self.max_volt_tolz = 0
        self.max_current_tolz = 0
        self.value_edit.returnPressed.connect(self.load_voltage_current)
##########################################################################################################################################################
        self.AC_DC_box.setEnabled(False)
        self.test_button.setEnabled(False)
        self.value_edit.setEnabled(False)
        self.connect_button.setEnabled(False)
        self.refresh_button.setEnabled(False)
        self.version_edit.setEnabled(False)
        self.firstMessage()
##########################################################################################################################################################

    def connect_multimeter(self):
        if not self.multimeter:
            try:
                self.multimeter = self.rm.open_resource('TCPIP0::192.168.222.207::INSTR')
                self.textBrowser.append(self.multimeter.query('*IDN?'))
                self.on_button_click('images_/images/PP7_3.jpg')
                self.start_button.setText('SPANNUNG')
                self.powersupply.write('OUTPut '+self.PS_channel+',ON')
                self.info_label.setText('Press SPANNUNG button and Check the Information')
            except visa.errors.VisaIOError:
                self.textBrowser.append('Multimeter has not been presented')
        else:
            self.multimeter.close()
            self.multimeter = None
            self.textBrowser.append(self.multimeter.query('*IDN?'))

    def connect_powersupply(self):
        if not self.powersupply:
            try:
                self.powersupply = self.rm.open_resource('TCPIP0::192.168.222.141::INSTR')
                self.textBrowser.setText(self.powersupply.resource_name)
                self.start_button.setEnabled(False)
                self.info_label.setText('write CH1 in the box next to CH and Press ENTER')
                self.value_edit.setEnabled(True)
                self.vals_button.setText('CH')
                self.on_button_click('images_/images/PP7.jpg')
            except visa.errors.VisaIOError:
                QMessageBox.information(self, "PowerSupply Connection", "PowerSupply is not present at the given IP Address.")
                self.textBrowser.setText('Powersupply has not been presented.')
        else:
            self.powersupply.close()
            self.powersupply = None
            self.PS_button.setText('PS ON')
            self.textBrowser.setText('Netzteil Disconnected')


    def firstMessage(self):
        msgBox = QMessageBox()
        msgBox.setWindowIcon(QIcon('images_/icons/icon.png'))
        msgBox.setText("Welcome to the testing world.")
        msgBox.setInformativeText("Press OK if you are ready.")
        msgBox.setWindowTitle("Message")
        self.on_button_click('images_/images/PP11.png')
        msgBox.setStandardButtons(QMessageBox.Ok)
        self.info_label.setText('  Press START Button  ')
        # msgBox.exec_()
        ret_value = msgBox.exec_()
        if ret_value == QMessageBox.Ok:
            self.secondMessage()

    def secondMessage(self):
        msgBox = QMessageBox()
        msgBox.setWindowIcon(QIcon('images_/icons/icon.png'))
        msgBox.setText("Press the PowerON buttons of PowerSupply and Multimeter to avoid delay.\n\nSet all the environment as shown in the image.\n\nRead the information carefully everytime.\n")
        msgBox.setInformativeText("Then press the button.")
        msgBox.setWindowTitle("Message")
        msgBox.setStandardButtons(QMessageBox.Ok)
        self.on_button_click('images_/images/PP1.jpg')
        self.title_label.setText('Preparation Test')
        self.info_label.setText('Press START Button')
        msgBox.exec_()

    def on_button_click(self, file_path):
        if file_path:
            pixmap = QPixmap(file_path)
            self.image_label.setPixmap(pixmap)
            self.image_label.setScaledContents(True)
            self.image_label.setFixedSize(pixmap.width(), pixmap.height())

            if self.start_button.text() == 'STEP4':
                reply = self.show_good_message()                
                if reply == QMessageBox.Yes:                    
                    self.start_button.setText('Netzteil ON')
                    self.on_button_click('images_/images/img4.jpg')
                    self.info_label.setText('Press Netzteil ON button')
                    
                else:
                    self.on_button_click('images_/icons/3.jpg')


            if self.start_button.text() == 'JUMPER CK':
                reply = self.jumper_close()
                if reply == QMessageBox.Yes:
                    self.start_button.setText('MULTI ON')
                    self.start_button.setEnabled(True)
                    self.on_button_click('images_/images/PP16.jpg')
                    self.info_label.setText('Press MULTI ON button')
                else:
                    self.on_button_click('images_/images/PP17.jpg')

    def show_good_message(self):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText("Good that you have correct Environment. Stage 1 has been Completed. Do you want to continue?")
        msgBox.setWindowTitle("Congratulations!")
        self.title_label.setText('Powersupply Test')
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        
        return msgBox.exec_()


    def connect(self):
        if self.start_button.text() == 'START':
            self.on_button_click('images_/images/PP4.jpg')
            self.info_label.setText("Place the board on ESD Matte See the Image on the right.\n \n Check all your environment with the image. Check All connections.\n \n Please close if anything not correct. If correct, press step1 Button.")
            self.start_button.setText("STEP1")
        elif self.start_button.text() == 'STEP1':
            self.on_button_click('images_/images/PP4_.jpg')
            self.info_label.setText('Check all the "4" screws are fitted according to the image.\n\n If not then please fit 4x "M2,5x5" Torx and M 2.5 Screws on X200 & X300 please press the button Step2.')
            self.start_button.setText("STEP2")
        elif self.start_button.text() == 'STEP2':
            self.on_button_click('images_/images/PP4.jpg')
            self.info_label.setText('Place back the Board on ESD Matte.')
            self.start_button.setText("STEP3")
        elif self.start_button.text()=='STEP3':
            self.on_button_click('images_/images/PP6.jpg')
            self.info_label.setText('Connect the other Cable to the board. as shown in the Image on the right.')
            self.start_button.setText("STEP4")
        elif self.start_button.text()=='STEP4':
            self.on_button_click('images_/icons/2.jpg')

        elif self.start_button.text()== 'Netzteil ON':
            self.connect_powersupply()

        elif self.start_button.text() == 'MESS V_I':
            self.calc_voltage_before_jumper()

        elif self.start_button.text() == 'JUMPER CK':
            self.on_button_click('images_/Power_ON_PS.jpg')

        elif self.start_button.text()== 'MULTI ON':
            self.connect_multimeter()

        elif self.start_button.text()== 'SPANNUNG':
            self.info_label.setText('Press STROM button')
            self.on_button_click('images_/images/PP7_4.jpg')
            self.start_button.setText('STROM')
            self.textBrowser.append(('PowerSUpply \n'+self.powersupply.query('MEASure:VOLTage? '+self.PS_channel)+'V'))

        elif self.start_button.text()== 'STROM':
            self.calc_voltage_before_jumper()

        elif (self.start_button.text()== 'Messung' and self.AC_DC_box.text == 'DCV'):
            self.mess_with_multimeter()
            self.start_button.setText('Messung')
        elif self.start_button.text()== 'Messung':
            self.mess_with_multimeter()

        else:
            self.start_button.setText("Start")
            self.textBrowser.append('disconnected')

    def mess_with_multimeter(self):
        if self.AC_DC_box.currentText() == 'DCV':
            result = self.multimeter.query('MEAS:VOLT:DC?')
            self.textBrowser.append(str(result))
            # return result.lstrip(">>")
            return result
        elif self.AC_DC_box.currentText()== 'ACV':
            result = self.multimeter.query('MEAS:VOLT:AC?')
            self.textBrowser.append(str(result))
            # return result.lstrip(">>")
            return result


    def handleSelectionChange(self, text):
        if text == 'DCV':
            self.selected_command = 'MEAS:VOLT:DC?'
            self.test_button.setEnabled(True)
            self.on_button_click('images_/images/R709_DC.jpg')
            self.test_button.setText('R709')
            self.info_label.setText('\n\n\n\n Press R709')
            self.AC_DC_box.setEnabled(False)

        elif text == 'ACV':
            self.selected_command = 'MEAS:VOLT:AC?'
            self.test_button.setEnabled(True)
            self.on_button_click('images_/pop.png')
            self.AC_DC_box.setEnabled(False)
            self.test_button.setText('R709')
            self.info_label.setText('Press R709 nutton.\ncheck voltage at R709 component \nshown in the picture.')
        else:
            self.selected_command = None
            self.test_button.setEnabled(False)
            self.info_label.setText('Select DC in the combobox to proceed further.')

    def on_cal_voltage_current(self):
        ret_volt = self.mess_with_multimeter()
        self.result_edit.setText(ret_volt)

        if self.AC_DC_box.currentText() == 'DCV' and self.test_button.text() == 'GO':
            self.multimeter.query('*IDN?')
            self.info_label.setText('See the Component in the Image and\n\n\n check the voltage at the same component.')          

        elif self.AC_DC_box.currentText() == 'DCV' and self.test_button.text() == 'R709':
            if 3.28 <= float(ret_volt) <= 3.38:
                self.textBrowser.append("DC Voltage at R709:"+ str(ret_volt))
                
            else:
                QMessageBox.information(self, "Status", "Voltage is diferred"+ret_volt)
            self.test_button.setText('R700')

        elif self.AC_DC_box.currentText() == 'DCV' and self.test_button.text() == 'R700':
            if 4.95 <= float(ret_volt) <= 5.05:
                self.textBrowser.append("DC Voltage at R700:"+ str(ret_volt))
            else:
                QMessageBox.information(self, "Status", "Voltage is diferred"+str(ret_volt))
            self.test_button.setEnabled(False)
            self.AC_DC_box.setEnabled(True)
            self.info_label.setText('Change the selection in AC/DC Box. \n\n\n Now we have to calculate AC Voltage at the last two components... \n\n\n So, Select ACV in AC/DC Box')
            QMessageBox.information(self, "Information", "Select ACV in the box near AC/DC")

        elif self.AC_DC_box.currentText() == 'ACV' and self.test_button.text() == 'R709':
            if float(ret_volt) <= 0.01:
                self.textBrowser.append("AC Voltage at R709:"+ str(ret_volt))
            else:
                QMessageBox.information(self, "Status", "Voltage is diferred"+str(ret_volt))
            self.test_button.setText('R700')

        elif self.AC_DC_box.currentText() == 'ACV' and self.test_button.text() == 'R700':
            if float(ret_volt) <= 0.01:
                self.textBrowser.append("AC Voltage at R700:"+ str(ret_volt))
            else:
                QMessageBox.information(self, "Status", "Voltage is diferred"+str(ret_volt))
            self.test_button.setEnabled(False)
            self.AC_DC_box.setEnabled(True)
            self.info_label.setText('Select DCV in AC DC Box')
            QMessageBox.information(self, "Information", "Select DCV in the box near AC/DC")


    def update_time_label(self):
        current_time = QTime.currentTime().toString(Qt.DefaultLocaleLongDate)
        current_date = QDate.currentDate().toString(Qt.DefaultLocaleLongDate)
        self.time_label.setText(f"{current_time} - {current_date}")
        # return current_date, current_time
    def update_com_ports(self, com_ports):
        self.port_box.clear()
        self.port_box.addItems(com_ports)


    def connect_or_disconnect_serial_port(self):
        if self.serial_port is None:
            com_port = self.port_box.currentText()            # Get the selected com port and baud rate
            baud_rate = int(self.baud_rates_combo.currentText())
            self.serial_port = serial.Serial(com_port, baud_rate, timeout=1)            # Create a new serial port object and open it
            self.port_box.setEnabled(False)  # Disable the combo boxes and change the button text
            self.baud_rates_combo.setEnabled(False)
            self.connect_button.setText('Disconnect')
            self.textBrowser.append('Serial Communication Connected')
            self.test_button.setEnabled(True)
            self.next_button.setEnabled(True)
            self.refresh_button.setEnabled(False)
        else:
            self.serial_port.close()            # Close the serial port
            self.serial_port = None
            self.connect_button.setEnabled(True)
            self.port_box.setEnabled(True)            # Enable the combo boxes and change the button text
            self.baud_rates_combo.setEnabled(True)
            self.refresh_button.setEnabled(True)
            self.next_button.setEnabled(False)
            self.test_button.setEnabled(False)
            self.connect_button.setText('Connect')
            self.textBrowser.append('Communication Disconnected')
    def refresh_connect(self):
        self.serial_thread.quit()
        self.serial_thread.wait()
        self.serial_thread.start()

    def start_process(self):
        if self.thread is None or not self.thread.isRunning():
            QMessageBox.information(self, "Process Started", "Process has been started.")
            self.thread = WorkerThread(self.commands, self.serial_port)
            self.thread.result_signal.connect(self.on_widget_button_clicked)
            self.thread.process_completed.connect(self.process_completed)
            self.thread.response_signal.connect(self.update_lineinsert)
            self.thread.start()
        else:
            QMessageBox.warning(self, "Process In Progress", "Process is already running.")
    def process_completed(self):
        QMessageBox.information(self, "Process Completed", "Process has been completed.")

    def load_voltage_current(self):
        if (self.vals_button.text() == 'CH' and self.value_edit.text() in ['ch1', 'CH1', 'Ch1', 'cH1']):
            self.powersupply.write('INSTrument CH1')
            self.PS_channel = self.value_edit.text()
            self.vals_button.setText('V')
            self.info_label.setText('Write 30 in the box next to V')
            self.value_edit.clear()
            self.on_button_click('images_/images/PP7_1.jpg')
            self.value_edit.setValidator(QRegExpValidator(QRegExp(r'^\d+(\.\d+)?$')))

        elif (self.vals_button.text() == 'CH' and self.value_edit.text() in ['ch2', 'CH2', 'Ch2', 'cH2']):
            self.powersupply.write('INSTrument CH2')
            self.PS_channel = self.value_edit.text()
            self.vals_button.setText('V')
            self.info_label.setText('Enter 30 in the box next to V')
            self.value_edit.clear()
            self.on_button_click('images_/images/PP7_1.jpg')
            self.value_edit.setValidator(QRegExpValidator(QRegExp(r'^\d+(\.\d+)?$')))

        elif (self.vals_button.text() == 'CH' and self.value_edit.text() in ['ch3', 'CH3', 'Ch3', 'cH3']):
            self.powersupply.write('INSTrument CH3')
            self.PS_channel = self.value_edit.text()
            self.vals_button.setText('V')
            self.info_label.setText('Enter 30 in the box next to V')
            self.value_edit.clear()
            self.on_button_click('images_/images/PP7_1.jpg')
            self.value_edit.setValidator(QRegExpValidator(QRegExp(r'^\d+(\.\d+)?$')))

        elif self.vals_button.text() == 'V':
            self.powersupply.write(self.PS_channel+':VOLTage ' + self.value_edit.text())
            self.textBrowser.append(self.powersupply.query(self.PS_channel+':VOLTage?'))
            self.vals_button.setText('I')
            self.info_label.setText('Enter 0.5 in the box next to I')
            self.value_edit.clear()
            self.on_button_click('images_/images/PP7_2.jpg')

        elif self.vals_button.text() == 'I':
            self.powersupply.write('CH1:CURRent ' + str(float(self.value_edit.text())))
            self.textBrowser.append(self.powersupply.query(self.PS_channel+':CURRent?'))
            self.vals_button.setText('Tolz V')
            self.info_label.setText('Enter 0.05 in the box next to I')
            self.powersupply.write('OUTPut '+self.PS_channel+',ON')
            self.value_edit.clear()
            self.on_button_click('images_/images/PP8.jpg')

        elif self.vals_button.text() == 'Tolz V':
            self.volt_toleranz = self.value_edit.text()
            self.textBrowser.append(self.volt_toleranz)
            self.value_edit.clear()
            self.info_label.setText('Enter 0.5 in the box next to I')
            self.vals_button.setText('Tolz I')
            self.on_button_click('images_/images/PP8_1.jpg')

        elif self.vals_button.text() == 'Tolz I':
            self.curr_toleranz = self.value_edit.text()
            self.textBrowser.append(self.curr_toleranz)
            self.value_edit.setEnabled(False)
            self.start_button.setEnabled(True)
            self.info_label.setText('Press MESS V_I')
            self.start_button.setText('MESS V_I')
            self.on_button_click('images_/images/PP17.jpg')
        else:
            self.textBrowser.append('Wrong Input')
    
    def calc_voltage_before_jumper(self):
        voltage = float(self.powersupply.query('MEASure:VOLTage? '+self.PS_channel))
        current = float(self.powersupply.query('MEASure:CURRent? '+self.PS_channel))
        self.result_edit.setText('Current: '+str(current))
        self.textBrowser.append('Volatge: '+str(voltage)+', Current: '+str(current))
        if self.start_button.text() == 'MESS V_I':
            if 0.04 <= current <= 0.06:
                self.start_button.setText('JUMPER CK')
                self.start_button.setEnabled(False)
            else:
                QMessageBox.information(self, 'Information', 'Supplying Current is either more or less. So please Swith OFF the PowerSupply, and Put back all the Euipment back.')
                self.powersupply.write('OUTPut '+self.PS_channel+',OFF')
                self.start_button.setText('JUMPER CK')
                self.start_button.setEnabled(False)
                self.info_label.setText('Close the JUMPER')
                self.on_button_click('images_/images/close_jumper.jpg')
        elif self.start_button.text() == 'STROM':
            if 0.09 <= current <= 0.15:
                self.start_button.setEnabled(False)
                self.AC_DC_box.setEnabled(True)
                self.test_button.setEnabled(True)
                self.on_button_click('images_/images/sel_DC_in_multimeter.jpg')
                QMessageBox.information(self, "Information", "Select DCV in the box near AC/DC")
                self.info_label.setText('\n \n \n \n Select DCV from AC/DC..!')
            else:
                QMessageBox.information(self, 'Information', 'Supplying Current is either more or less. So please Swith OFF the PowerSupply, and Put back all the Euipment back.')
                self.powersupply.write('OUTPut '+self.PS_channel+',OFF')
                
    def jumper_close(self):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText("Close the Jumper Before proceed further. If closed the Press Yes")
        msgBox.setWindowTitle("IMPORTANT!")
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        return msgBox.exec_()

##########################################################################################################            
    def Load_Excel(self):
        file_dialog = QFileDialog(self)        # Open file dialog to select an Excel sheet
        file_dialog.setNameFilter("Excel Files (*.xlsx *.xls)")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            file_path = selected_files[0]
        else:
            return
        try:
            excel_data = pd.read_excel(file_path)        # Load the selected Excel sheet
        except pd.errors.EmptyDataError:
            QMessageBox.warning(self, "Empty Sheet", "The selected Excel sheet is empty.")
            return
        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred while loading the Excel sheet: {str(e)}")
            return
        self.textBrowser.clear()        # Display the data in the text browser
        self.textBrowser.append("Loaded Excel Data:" + ' ')
        self.textBrowser.append(excel_data.to_string())
                                                                                                ###############################################################################
    def Save_Excel(self):
        port = self.com_ports_combo.currentText()        # Get data from form fields
        baud = self.baud_rates_combo.currentText()
        DC_Vals = self.DC_AC_Values[0][1]
        AC_Vals = self.DC_AC_Values[1][2]

        name, ok = QInputDialog.getText(self, "Save Data", "Enter a name:")        # Ask the user for the name using a popup dialog
        if not ok:
            return
        workbook_name = "mydata.xlsx"        # Open or create Excel workbook and worksheet
        if os.path.exists(workbook_name):
            workbook = openpyxl.load_workbook(workbook_name)
        else:
            workbook = openpyxl.Workbook()
        worksheet = workbook.active
        name_exists = False        # Check if name already exists in the sheet
        for row in worksheet.iter_rows(values_only=True):
            if name == row[0]:
                name_exists = True
                break
        if name_exists:
            reply = QMessageBox.question(
                self, "Name already exists",
                f"{name} already exists. Do you want to save with a different name?",
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                new_name, ok = QInputDialog.getText(self, "Save Data", "Enter a different name:")                # Ask the user for a different name using a popup dialog
                if not ok:
                    return
                name = new_name
        data = [name, port, baud, DC_Vals, AC_Vals]        # Add data to sheet  # Add more data here
        worksheet.append(data)
        workbook.save(workbook_name)        # Save the workbook
        # Optionally, you can open the saved workbook in an external application
        # subprocess.run(["open", workbook_name])
#####################################################################################################################################################################################
    def closeEvent(self, event):
        if not self.powersupply is None:
            try:
                response = self.powersupply.query(f"OUTP? CH{self.PS_channel}")
                print(response)
                if not response == None:
                    self.powersupply.write('OUTPut '+self.PS_channel+',OFF')
                else:
                    pass
            except visa.errors.InvalidSession:
                print('Error in PS')
        else:
            pass
        event.accept()

    def Load_Excel(self):
        file_dialog = QFileDialog(self)        # Open file dialog to select an Excel sheet
        file_dialog.setNameFilter("Excel Files (*.xlsx *.xls)")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            file_path = selected_files[0]
        else:
            return
        try:
            excel_data = pd.read_excel(file_path)        # Load the selected Excel sheet
        except pd.errors.EmptyDataError:
            QMessageBox.warning(self, "Empty Sheet", "The selected Excel sheet is empty.")
            return
        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred while loading the Excel sheet:\n{str(e)}")
            return
        self.textBrowser.clear()        # Display the data in the text browser
        self.textBrowser.append("Loaded Excel Data:" + '\n')
        self.textBrowser.append(excel_data.to_string())
#####################################################################################################################################################################################
    def Save_Excel(self):
        port = self.port_box.currentText()        # Get data from form fields
        baud = self.baudrate_box.currentText()
        # DC_Vals = self.DC_AC_Values[0][1]
        # AC_Vals = self.DC_AC_Values[1][2]

        name, ok = QInputDialog.getText(self, "Save Data", "Enter a name:")        # Ask the user for the name using a popup dialog
        if not ok:
            return
        workbook_name = "mydata.xlsx"        # Open or create Excel workbook and worksheet
        if os.path.exists(workbook_name):
            workbook = openpyxl.load_workbook(workbook_name)
        else:
            workbook = openpyxl.Workbook()
        worksheet = workbook.active
        name_exists = False        # Check if name already exists in the sheet
        for row in worksheet.iter_rows(values_only=True):
            if name == row[0]:
                name_exists = True
                break
        if name_exists:
            reply = QMessageBox.question(
                self, "Name already exists",
                f"{name} already exists. Do you want to save with a different name?",
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                new_name, ok = QInputDialog.getText(self, "Save Data", "Enter a different name:")                # Ask the user for a different name using a popup dialog
                if not ok:
                    return
                name = new_name
        data = [name, port, baud]#, DC_Vals, AC_Vals]        # Add data to sheet  # Add more data here
        worksheet.append(data)
        workbook.save(workbook_name)        # Save the workbook
        # Optionally, you can open the saved workbook in an external application
        # subprocess.run(["open", workbook_name])


def main():
    app = QApplication(sys.argv)
    Window = App()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
